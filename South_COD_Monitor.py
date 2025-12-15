"""
South COD Monitor
Script to extract data from Google Sheets and generate reports for all hubs
"""

import pandas as pd
import numpy as np
import gspread
from google.oauth2.service_account import Credentials
from gspread_dataframe import set_with_dataframe
from datetime import datetime
from dateutil import parser as date_parser
import os
import sys
import time
import string
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Google Sheets Configuration
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

# Service account file for authentication
SERVICE_ACCOUNT_FILE = 'service_account_key.json'

# Google Sheets URL: https://docs.google.com/spreadsheets/d/1t04OxK-GdiDDUq85HNKtyDO2GqYDBoX2eG0M34aR3jA/edit?gid=1176716609#gid=1176716609
SPREADSHEET_ID = '1t04OxK-GdiDDUq85HNKtyDO2GqYDBoX2eG0M34aR3jA'
SOURCE_WORKSHEET_NAME = 'Dashboard'  # Worksheet name to extract data from

# Output Google Sheet (where to push the data)
OUTPUT_SPREADSHEET_ID = '1FUH-Z98GFcCTIKpSAeZPGsjIESMVgBB2vrb6QOZO8mM'
OUTPUT_WORKSHEET_NAME = 'COD Monitor'

# Columns to preserve in destination sheet (manually editable, should not be overwritten)
PRESERVE_COLUMNS = ['Van Adhoc', 'Legal Issue', 'Old Balance']

# Calculated column
CALCULATED_COLUMN = 'Actual Gap'  # Actual Gap = Overall Gap - (Van Adhoc + Legal Issue + Old Balance + Gap Date)

# CLM Email Mapping (from G-Form_COD_Status.py)
CLM_EMAIL = {
    "Asif": "abdulasif@loadshare.net",
    "Kishore": "kishorkumar.m@loadshare.net",
    "Haseem": "hasheem@loadshare.net",
    "Madvesh": "madvesh@loadshare.net",
    "Irappa": "irappa.vaggappanavar@loadshare.net",
    "Lokesh": "lokeshh@loadshare.net",
    "Bharath": "bharath.s@loadshare.net"
}

# Hub Email Mapping (from G-Form_COD_Status.py)
HUB_EMAIL = {
    "BagaluruMDH_BAG": "bagalurumdh_bag@loadshare.net",
    "KoorieeSoukyaRdODH_BLR": "koorieesoukyardodh_blr@loadshare.net",
    "KoorieeSoukyaRdTempODH_BLR": "soukyaodh_blr@loadshare.net",
    "SulebeleMDH_SUL": "sulebelemdh_sul@loadshare.net",
    "DommasandraSplitODH_DMN": "dommasandrasplitodh_dmn@loadshare.net",
    "HulimavuHub_BLR": "hulimavuhub_blr@loadshare.net",
    "ThavarekereMDH_THK": "thavarekeremdh_thk@loadshare.net",
    "KoorieeHayathnagarODH_HYD": "koorieeodh_hyd@loadshare.net",
    "SaidabadSplitODH_HYD": "saidabad.spiltodh@loadshare.net",
    "CABTSRNagarODH_HYD": "cabtsrnagarodh_hyd@loadshare.net",
    "LargeLogicKuniyamuthurODH_CJB": "kuniyamuthurodh_cjb@loadshare.net",
    "TTSPLKodaikanalODH_KDI": "kodaikanalODH_KDI@loadshare.net",
    "LargeLogicDharapuramODH_DHP": "dharapuramodh_dhp@loadshare.net",
    "TTSPLBatlagunduODH_BGU": "batlagunduODH_BGU@loadshare.net",
    "VadipattiMDH_VDP": "vadipattimdh_vdp@loadshare.net",
    "SITICSWadiODH_WDI": "wadiodh@loadshare.net",
    "BidarFortHub_BDR": "bidarfort_bdr@loadshare.net",
    "ElasticRunBidarODH_BDR": "bidarodh_bdr@loadshare.net",
    "NaubadMDH_BDR": "naubadmdh_brr@loadshare.net",
    "LargeLogicRameswaramODH_RMS": "rameswaramodh_rms@loadshare.net",
    "LargelogicChinnamanurODH_CNM": "chinnamannur_cnm@loadshare.net",
}

def get_email_recipients():
    """Build email recipient lists dynamically (from G-Form_COD_Status.py)"""
    # Get hub emails for hubs in HUBS list
    hub_emails = []
    for hub in HUBS:
        if hub in HUB_EMAIL:
            hub_emails.append(HUB_EMAIL[hub])
    
    # Get all CLM emails (handle multiple CLMs per hub)
    clm_emails = set()
    for hub, clm_info in HUB_INFO.items():
        clm_names = clm_info[0]  # First element is CLM name(s)
        # Handle multiple CLMs separated by comma
        clm_list = [name.strip() for name in clm_names.split(',')]
        for clm_name in clm_list:
            if clm_name in CLM_EMAIL:
                clm_emails.add(CLM_EMAIL[clm_name])
    
    # Add Lokesh, Bharath, and Maligai Rasmeen
    additional_recipients = [
        CLM_EMAIL.get("Lokesh", "lokeshh@loadshare.net"),
        CLM_EMAIL.get("Bharath", "bharath.s@loadshare.net"),
        "maligai.rasmeen@loadshare.net"
    ]
    
    # Combine all TO recipients (hubs + CLMs + Lokesh + Bharath + Maligai Rasmeen)
    to_recipients = list(set(hub_emails + list(clm_emails) + additional_recipients))
    
    # BCC list: Rakib only
    bcc_recipients = ["rakib@loadshare.net"]
    
    # CC list: Empty (no CC recipients)
    cc_recipients = []
    
    return to_recipients, cc_recipients, bcc_recipients

# Email Configuration (same as reservations_email_automation.py)
EMAIL_CONFIG = {
    'sender_email': os.getenv('GMAIL_SENDER_EMAIL', 'arunraj@loadshare.net'),
    'sender_password': os.getenv('GMAIL_APP_PASSWORD', 'ihczkvucdsayzrsu'),
    'recipient_email': 'arunraj@loadshare.net',  # Will be overridden by get_email_recipients()
    'cc_list': ['maligai.rasmeen@loadshare.net', 'rakib@loadshare.net'],  # Will be overridden by get_email_recipients()
    'smtp_server': 'smtp.gmail.com',
    'smtp_port': 587
}
EMAIL_ENABLED = True  # Set to False to disable email sending
TEST_MODE = False  # Set to True to mute recipients and send to test email only
TEST_EMAIL = 'arunraj@loadshare.net'  # Test email address (usually sender's own email)

# Output file name
OUTPUT_FILE = 'South_COD_Monitor_Report.xlsx'

# Columns to extract (case-insensitive matching)
COLUMNS_TO_EXTRACT = [
    'Total Collection',
    'Total Deposit',
    'Overall Gap'
]

# Columns to extract for latest date (date is in row 1, headers in row 2)
LATEST_DATE_COLUMNS = ['Collection', 'Gap']

# ============================================================================
# HUB CONFIGURATION (Permanent list - no need to refer other scripts)
# ============================================================================

# List of all hubs (21 hubs total)
HUBS = [
    "BagaluruMDH_BAG",
    "KoorieeSoukyaRdODH_BLR",
    "KoorieeSoukyaRdTempODH_BLR",
    "SulebeleMDH_SUL",
    "DommasandraSplitODH_DMN",
    "HulimavuHub_BLR",
    "ThavarekereMDH_THK",
    "KoorieeHayathnagarODH_HYD",
    "SaidabadSplitODH_HYD",
    "CABTSRNagarODH_HYD",
    "LargeLogicKuniyamuthurODH_CJB",
    "TTSPLKodaikanalODH_KDI",
    "LargeLogicDharapuramODH_DHP",
    "TTSPLBatlagunduODH_BGU",
    "VadipattiMDH_VDP",
    "SITICSWadiODH_WDI",
    "BidarFortHub_BDR",
    "ElasticRunBidarODH_BDR",
    "NaubadMDH_BDR",
    "LargeLogicRameswaramODH_RMS",
    "LargelogicChinnamanurODH_CNM"
]

# Mapping of hub to CLM Name and State
HUB_INFO = {
    "BagaluruMDH_BAG": ("Kishore", "Karnataka"),
    "NaubadMDH_BDR": ("Haseem", "Karnataka"),
    "SITICSWadiODH_WDI": ("Haseem", "Karnataka"),
    "VadipattiMDH_VDP": ("Madvesh", "Tamil Nadu"),
    "TTSPLKodaikanalODH_KDI": ("Madvesh", "Tamil Nadu"),
    "LargeLogicRameswaramODH_RMS": ("Madvesh", "Tamil Nadu"),
    "CABTSRNagarODH_HYD": ("Asif, Haseem", "Telengana"),
    "LargeLogicKuniyamuthurODH_CJB": ("Madvesh", "Tamil Nadu"),
    "KoorieeHayathnagarODH_HYD": ("Asif, Haseem", "Telengana"),
    "SulebeleMDH_SUL": ("Kishore", "Karnataka"),
    "KoorieeSoukyaRdODH_BLR": ("Kishore", "Karnataka"),
    "KoorieeSoukyaRdTempODH_BLR": ("Kishore", "Karnataka"),
    "ThavarekereMDH_THK": ("Irappa", "Karnataka"),
    "SaidabadSplitODH_HYD": ("Asif, Haseem", "Telengana"),
    "LargelogicChinnamanurODH_CNM": ("Madvesh", "Tamil Nadu"),
    "LargeLogicDharapuramODH_DHP": ("Madvesh", "Tamil Nadu"),
    "HulimavuHub_BLR": ("Kishore", "Karnataka"),
    "ElasticRunBidarODH_BDR": ("Haseem", "Karnataka"),
    "DommasandraSplitODH_DMN": ("Kishore", "Karnataka"),
    "TTSPLBatlagunduODH_BGU": ("Madvesh", "Tamil Nadu"),
    "BidarFortHub_BDR": ("Haseem", "Karnataka"),
}


def format_date_for_column(date_str):
    """Format date string to DD-MMM format (e.g., '11-Dec', '31-Dec')"""
    try:
        if not date_str or not str(date_str).strip():
            return str(date_str).strip() if date_str else "Latest"
        
        date_str_clean = str(date_str).strip()
        
        # Try to parse the date
        try:
            # Use current year as default if year is not specified
            current_year = datetime.now().year
            parsed_date = date_parser.parse(date_str_clean, fuzzy=True, default=datetime(current_year, 1, 1))
            
            # Format as DD-MMM (e.g., 11-Dec, 31-Dec)
            formatted_date = parsed_date.strftime('%d-%b')
            
            return formatted_date
        except:
            # If parsing fails, try to extract day and month manually
            # Handle formats like "11-December", "31-December", etc.
            parts = date_str_clean.split('-')
            if len(parts) >= 2:
                day = parts[0].strip()
                month_str = parts[1].strip()
                
                # Map full month names to abbreviations
                month_map = {
                    'january': 'Jan', 'february': 'Feb', 'march': 'Mar',
                    'april': 'Apr', 'may': 'May', 'june': 'Jun',
                    'july': 'Jul', 'august': 'Aug', 'september': 'Sep',
                    'october': 'Oct', 'november': 'Nov', 'december': 'Dec'
                }
                
                month_lower = month_str.lower()
                if month_lower in month_map:
                    return f"{day}-{month_map[month_lower]}"
                elif len(month_str) >= 3:
                    # Already abbreviated or short form
                    return f"{day}-{month_str[:3].capitalize()}"
                else:
                    return f"{day}-{month_str}"
            
            # If all parsing fails, return original string
            return date_str_clean
    except Exception as e:
        # If any error occurs, return original string
        return str(date_str).strip() if date_str else "Latest"


def setup_google_sheets():
    """Setup Google Sheets connection using service account"""
    try:
        if not os.path.exists(SERVICE_ACCOUNT_FILE):
            print(f"❌ Error: Service account key file not found: {SERVICE_ACCOUNT_FILE}")
            print("Please ensure the service_account_key.json file is in the same directory.")
            sys.exit(1)
        
        print("🔑 Setting up Google Sheets connection...")
        credentials = Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE,
            scopes=SCOPES
        )
        client = gspread.authorize(credentials)
        print("✅ Google Sheets connection established")
        
        # Display service account email for sharing reference
        try:
            import json
            with open(SERVICE_ACCOUNT_FILE, 'r') as f:
                service_account_data = json.load(f)
                service_account_email = service_account_data.get('client_email', 'Not found')
                print(f"📧 Service Account Email: {service_account_email}")
                print("💡 Make sure the Google Sheet is shared with this email address (Editor access)")
        except Exception as e:
            print(f"⚠️ Could not extract service account email: {e}")
        
        return client
    except Exception as e:
        print(f"❌ Error setting up Google Sheets: {e}")
        raise


def get_worksheet_by_name(client, spreadsheet_id, worksheet_name):
    """Get worksheet by name"""
    try:
        spreadsheet = client.open_by_key(spreadsheet_id)
        print(f"✅ Opened spreadsheet: {spreadsheet.title}")
        
        # List all worksheets
        worksheets = spreadsheet.worksheets()
        print(f"\n📊 Available worksheets ({len(worksheets)}):")
        for i, ws in enumerate(worksheets, 1):
            print(f"   {i}. {ws.title} (ID: {ws.id})")
        
        # Find worksheet by name (case-insensitive)
        target_worksheet = None
        for ws in worksheets:
            if ws.title.lower() == worksheet_name.lower():
                target_worksheet = ws
                print(f"\n✅ Found target worksheet: '{ws.title}'")
                break
        
        if not target_worksheet:
            print(f"\n⚠️ Warning: Could not find worksheet named '{worksheet_name}'")
            print("Available worksheets:")
            for ws in worksheets:
                print(f"   - {ws.title}")
            print("\nTrying to use the first worksheet instead...")
            target_worksheet = worksheets[0] if worksheets else None
            if target_worksheet:
                print(f"⚠️ Using '{target_worksheet.title}' as fallback")
        
        return target_worksheet
    except Exception as e:
        print(f"❌ Error accessing spreadsheet: {e}")
        raise


def find_hub_column(df):
    """Find the column that contains hub names"""
    hub_column = None
    
    # Common column names for hub
    possible_names = ['hub', 'hub name', 'hub_name', 'station', 'station name', 'station_name', 
                     'location', 'hub code', 'hub_code', 'name']
    
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if col_lower in possible_names:
            hub_column = col
            break
    
    # If not found, check if any column contains hub names from our list
    if not hub_column:
        for col in df.columns:
            try:
                # Ensure we get a Series, not DataFrame
                col_data = df[col]
                if isinstance(col_data, pd.DataFrame):
                    # If it's a DataFrame, take the first column
                    col_data = col_data.iloc[:, 0]
                
                # Check first few non-null values
                sample_values = col_data.dropna().head(10).astype(str).str.strip()
                for hub in HUBS:
                    if any(hub.lower() in str(val).lower() for val in sample_values):
                        hub_column = col
                        print(f"✅ Found hub column by matching hub names: '{col}'")
                        break
                if hub_column:
                    break
            except Exception as e:
                # Skip columns that cause errors
                continue
    
    return hub_column


def find_columns_to_extract(df):
    """Find the columns that match the required column names (case-insensitive)"""
    found_columns = {}
    
    print(f"\n🔍 Searching for required columns...")
    
    for required_col in COLUMNS_TO_EXTRACT:
        found = False
        required_lower = required_col.lower().strip()
        
        for col in df.columns:
            col_lower = str(col).lower().strip()
            # Exact match or contains the key words
            if col_lower == required_lower or required_lower in col_lower:
                found_columns[required_col] = col
                print(f"   ✅ Found '{required_col}' -> '{col}'")
                found = True
                break
        
        if not found:
            print(f"   ⚠️ Column '{required_col}' not found")
    
    return found_columns


def find_latest_date_columns(all_values, headers):
    """Find Collection and Gap columns by finding the last column with updated values"""
    try:
        if len(all_values) < 3:
            return {}, None
        
        # Row 1 contains dates (index 0)
        date_row = all_values[0]
        # Row 2 contains headers (index 1)
        header_row = all_values[1]
        # Data starts from row 3 (index 2)
        data_rows = all_values[2:]
        
        print(f"\n📅 Finding last column with updated values...")
        
        # Find the rightmost column group (Collection, Deposit, Gap) that has data
        # Look for columns that have non-zero, non-empty values
        num_cols = len(header_row)
        column_has_data = {}
        
        # Check each column for data (skip first few columns which are likely Hub Name, etc.)
        # Start from a reasonable column index (skip Hub Name and other fixed columns)
        start_check_col = 5  # Skip first few columns
        
        for col_idx in range(start_check_col, num_cols):
            has_data = False
            non_zero_count = 0
            
            # Check data rows for this column
            for row in data_rows:
                if col_idx < len(row):
                    cell_value = str(row[col_idx]).strip() if row[col_idx] else ''
                    # Check if it's a number (currency format or plain number)
                    if cell_value:
                        # Remove currency symbols and commas
                        clean_value = cell_value.replace('₹', '').replace(',', '').replace(' ', '').strip()
                        try:
                            num_value = float(clean_value)
                            if num_value != 0:
                                has_data = True
                                non_zero_count += 1
                        except:
                            pass
            
            if has_data:
                column_has_data[col_idx] = non_zero_count
        
        if not column_has_data:
            print("   ⚠️ No columns with data found")
            return {}, None
        
        # Find the rightmost column with data
        rightmost_col_idx = max(column_has_data.keys())
        print(f"   ✅ Rightmost column with data: Column {rightmost_col_idx} ({column_has_data[rightmost_col_idx]} non-zero values)")
        
        # Get the date from row 1 for this column (or nearby columns if merged)
        latest_date_str = None
        # Check current column and previous columns for date (merged headers)
        for offset in [0, -1, -2]:
            check_idx = rightmost_col_idx + offset
            if check_idx >= 0 and check_idx < len(date_row):
                date_value = date_row[check_idx] if check_idx < len(date_row) else ''
                if date_value and str(date_value).strip():
                    latest_date_str = str(date_value).strip()
                    print(f"   📅 Found date '{latest_date_str}' at column {check_idx}")
                    break
        
        if not latest_date_str:
            # Try to find any date in row 1 near the rightmost column
            for check_idx in range(max(0, rightmost_col_idx - 5), min(len(date_row), rightmost_col_idx + 1)):
                date_value = date_row[check_idx] if check_idx < len(date_row) else ''
                if date_value and str(date_value).strip():
                    latest_date_str = str(date_value).strip()
                    print(f"   📅 Found date '{latest_date_str}' at column {check_idx} (near rightmost)")
                    break
        
        if not latest_date_str:
            latest_date_str = "Latest"
            print(f"   ⚠️ Could not find date, using 'Latest'")
        
        # For merged headers, Collection, Deposit, Gap are typically the last 3 columns
        # Collection is first (idx-2), Deposit is middle (idx-1), Gap is last (idx)
        # Find the range: typically 3 columns ending at rightmost_col_idx
        latest_date_range = []
        if rightmost_col_idx >= 2:
            # Assume merged range is 3 columns: Collection (idx-2), Deposit (idx-1), Gap (idx)
            latest_date_range = [rightmost_col_idx - 2, rightmost_col_idx - 1, rightmost_col_idx]
        elif rightmost_col_idx >= 1:
            latest_date_range = [rightmost_col_idx - 1, rightmost_col_idx]
        else:
            latest_date_range = [rightmost_col_idx]
        
        print(f"   📍 Column range for latest data: {latest_date_range} (Collection at {latest_date_range[0]}, Gap at {latest_date_range[-1]})")
        
        # Now look for Collection and Gap headers in row 2 (headers row)
        # Handle merged cells: if header is empty, check previous non-empty header
        found_columns = {}
        
        # Build a mapping of column index to header (handling merged cells)
        header_map = {}
        last_non_empty_header = None
        for col_idx in range(len(header_row)):
            header_value = header_row[col_idx] if col_idx < len(header_row) else ''
            if header_value and str(header_value).strip():
                last_non_empty_header = str(header_value).strip()
                header_map[col_idx] = last_non_empty_header
            elif last_non_empty_header:
                # Merged cell - use previous header
                header_map[col_idx] = last_non_empty_header
        
        print(f"   📋 Header mapping (handling merged cells) - last 15 columns:")
        for col_idx in sorted(header_map.keys())[-15:]:  # Show last 15 for debugging
            print(f"      Column {col_idx}: '{header_map[col_idx]}'")
        
        # Find Collection and Gap columns for the latest date
        # Look in the date range for these specific headers
        # For merged dates, Collection is typically the first column, Gap is the last (3rd) column
        for col_name in LATEST_DATE_COLUMNS:
            col_name_lower = col_name.lower().strip()
            found = False
            
            # First, search within the latest date range
            # For merged headers: Collection is first column, Gap is last (3rd) column
            if latest_date_range:
                print(f"   🔍 Searching for '{col_name}' in date range: {latest_date_range}")
                
                # For merged columns: Collection is first, Gap is last
                if col_name_lower == 'collection':
                    # Collection is the first column in the merged range
                    target_idx = latest_date_range[0]
                    print(f"      → Targeting first column in range: index {target_idx}")
                elif col_name_lower == 'gap':
                    # Gap is the last column in the merged range (should be 3rd column)
                    target_idx = latest_date_range[-1] if len(latest_date_range) >= 3 else latest_date_range[-1]
                    print(f"      → Targeting last column in range: index {target_idx}")
                else:
                    target_idx = None
                
                # Verify the header matches at target index
                if target_idx is not None:
                    if target_idx in header_map:
                        header = header_map[target_idx]
                        header_lower = str(header).lower().strip()
                        print(f"      → Header at index {target_idx}: '{header}'")
                        
                        # Check if header matches
                        if col_name_lower == header_lower or col_name_lower in header_lower:
                            found_columns[col_name] = {
                                'column_index': target_idx,
                                'header': header,
                                'date': latest_date_str
                            }
                            print(f"   ✅ Found '{col_name}' column at index {target_idx}: '{header}' (Date: {latest_date_str})")
                            found = True
                        else:
                            # Header doesn't match, search within the range
                            print(f"      → Header doesn't match, searching within range...")
                            for col_idx in latest_date_range:
                                if col_idx in header_map:
                                    header = header_map[col_idx]
                                    header_lower = str(header).lower().strip()
                                    print(f"         Checking index {col_idx}: '{header}'")
                                    if col_name_lower == header_lower or col_name_lower in header_lower:
                                        found_columns[col_name] = {
                                            'column_index': col_idx,
                                            'header': header,
                                            'date': latest_date_str
                                        }
                                        print(f"   ✅ Found '{col_name}' column at index {col_idx}: '{header}' (Date: {latest_date_str})")
                                        found = True
                                        break
                    else:
                        print(f"      ⚠️ Index {target_idx} not in header_map")
            
            # If not found in date range, search all columns (fallback)
            if not found:
                for col_idx, header in header_map.items():
                    header_lower = str(header).lower().strip()
                    if col_name_lower == header_lower or col_name_lower in header_lower:
                        # Check if this column is in or near the latest date range
                        if col_idx in latest_date_range or (latest_date_range and abs(col_idx - latest_date_range[0]) <= 3):
                            found_columns[col_name] = {
                                'column_index': col_idx,
                                'header': header,
                                'date': latest_date_str
                            }
                            print(f"   ✅ Found '{col_name}' column at index {col_idx}: '{header}' (Date: {latest_date_str}) [fallback]")
                            found = True
                            break
        
        if len(found_columns) < len(LATEST_DATE_COLUMNS):
            missing = [col for col in LATEST_DATE_COLUMNS if col not in found_columns]
            print(f"   ⚠️ Could not find columns: {missing}")
        
        return found_columns, latest_date_str
        
    except Exception as e:
        print(f"   ⚠️ Error finding latest date columns: {e}")
        import traceback
        traceback.print_exc()
        return {}, None


def extract_sheet_data(worksheet):
    """Extract data from the worksheet, filtered by hub names"""
    try:
        print(f"\n📥 Reading data from worksheet: '{worksheet.title}'...")
        
        # Get all values from the worksheet
        all_values = worksheet.get_all_values()
        
        if not all_values or len(all_values) < 2:
            print("⚠️ Warning: No data found in the worksheet or insufficient rows")
            return pd.DataFrame()
        
        # Headers are in row 2 (index 1)
        headers = all_values[1]  # Row 2 (0-indexed, so index 1)
        print(f"📋 Found {len(headers)} columns (headers from row 2)")
        
        # Get data rows starting from row 3 (index 2)
        data_rows = all_values[2:] if len(all_values) > 2 else []
        print(f"📊 Found {len(data_rows)} data rows (starting from row 3)")
        
        # Find Collection and Gap columns for latest date (date is in row 1, headers in row 2)
        latest_date_columns, latest_date_str = find_latest_date_columns(all_values, headers)
        
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=headers)
        
        # Handle duplicate column names by making them unique
        if df.columns.duplicated().any():
            print("⚠️ Warning: Found duplicate column names. Making them unique...")
            df.columns = [f"{col}_{i}" if col in df.columns[:i].tolist() else col 
                         for i, col in enumerate(df.columns)]
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        print(f"✅ Successfully extracted {len(df)} rows of data")
        print(f"\n📋 Column names:")
        for i, col in enumerate(df.columns, 1):
            print(f"   {i}. {col}")
        
        # Find the hub column - specifically look for "Hub Name"
        hub_column = None
        for col in df.columns:
            if str(col).strip().lower() == 'hub name':
                hub_column = col
                print(f"\n✅ Found 'Hub Name' column: '{col}'")
                break
        
        # If not found, try the find_hub_column function as fallback
        if not hub_column:
            hub_column = find_hub_column(df)
            if hub_column:
                print(f"   Hub column identified by pattern matching: '{hub_column}'")
        
        if hub_column:
            print(f"\n🔍 Filtering data for {len(HUBS)} specific hubs...")
            print(f"   Hub column: '{hub_column}'")
            
            # Filter rows that exactly match or contain any of our 21 hubs
            # First, ensure we have a Series
            hub_series = df[hub_column]
            if isinstance(hub_series, pd.DataFrame):
                hub_series = hub_series.iloc[:, 0]
            
            # Create a mask for rows that match any of our 21 hubs
            def matches_hub(hub_value):
                if pd.isna(hub_value):
                    return False
                hub_str = str(hub_value).strip()
                for hub in HUBS:
                    # Exact match or hub name contained in the value
                    if hub.lower() == hub_str.lower() or hub.lower() in hub_str.lower():
                        return True
                return False
            
            mask = hub_series.apply(matches_hub)
            
            filtered_df = df[mask].copy()
            
            print(f"   Found {len(filtered_df)} rows matching hub names")
            print(f"   Filtered out {len(df) - len(filtered_df)} rows")
            
            # Find the required columns
            found_columns = find_columns_to_extract(filtered_df)
            
            # Build the final DataFrame with only required columns
            columns_to_keep = [hub_column]  # Always include hub column
            
            # Add the found columns (using their actual names from the sheet)
            for required_col, actual_col in found_columns.items():
                if actual_col not in columns_to_keep:
                    columns_to_keep.append(actual_col)
            
            # Add latest date columns (Collection and Gap)
            latest_date_column_mapping = {}
            if latest_date_columns:
                print(f"\n📅 Adding latest date columns...")
                for col_name, col_info in latest_date_columns.items():
                    col_idx = col_info['column_index']
                    original_header = col_info['header']
                    
                    # Check if column exists in DataFrame by index
                    if col_idx < len(filtered_df.columns):
                        # Get the actual column name from DataFrame (may differ due to duplicate handling)
                        actual_col_name = filtered_df.columns[col_idx]
                        
                        if actual_col_name not in columns_to_keep:
                            columns_to_keep.append(actual_col_name)
                        
                        # Create destination column name: "Colc Date" or "Gap Date"
                        # Use "Colc" instead of "Collection" and format date as DD-MMM (e.g., 11-Dec)
                        display_name = "Colc" if col_name.lower() == "collection" else col_name
                        formatted_date = format_date_for_column(col_info['date'])
                        dest_col_name = f"{display_name} {formatted_date}"
                        latest_date_column_mapping[actual_col_name] = dest_col_name
                        print(f"   ✅ Column {col_idx}: '{original_header}' -> '{dest_col_name}'")
                    else:
                        print(f"   ⚠️ Column index {col_idx} out of range for '{col_name}'")
            else:
                print(f"\n⚠️ No latest date columns found to add")
            
            # Create final DataFrame with only selected columns
            final_df = filtered_df[columns_to_keep].copy()
            
            # Rename columns to standard names if they differ
            column_mapping = {}
            for required_col, actual_col in found_columns.items():
                if actual_col != required_col:
                    column_mapping[actual_col] = required_col
            
            # Add latest date column mappings
            column_mapping.update(latest_date_column_mapping)
            
            if column_mapping:
                final_df = final_df.rename(columns=column_mapping)
                print(f"\n📝 Renamed columns for consistency:")
                for old_name, new_name in column_mapping.items():
                    print(f"   '{old_name}' -> '{new_name}'")
            
            # Reorder columns: Move Colc Date and Gap Date to the end (they should be the LAST two columns)
            if latest_date_column_mapping:
                print(f"\n📋 Reordering columns to place latest date columns at the end...")
                current_columns = list(final_df.columns)
                
                # Find the latest date column names (after renaming)
                latest_date_col_names = list(latest_date_column_mapping.values())
                
                # Remove latest date columns from current position
                remaining_columns = [col for col in current_columns if col not in latest_date_col_names]
                
                # Always place latest date columns at the very end (after all other columns including Actual Gap)
                reordered_columns = remaining_columns + latest_date_col_names
                
                # Reorder the DataFrame
                final_df = final_df[reordered_columns]
                print(f"   ✅ Columns reordered. Latest date columns placed at the end: {latest_date_col_names}")
                print(f"   📋 Final column order: {list(final_df.columns)}")
            
            # Show which hubs were found (hub column won't be renamed, so use original name)
            found_hubs = final_df[hub_column].unique() if hub_column in final_df.columns else []
            print(f"\n📌 Hubs found in data ({len(found_hubs)}):")
            for hub in sorted(found_hubs):
                print(f"   ✓ {hub}")
            
            # Show which hubs are missing
            if len(found_hubs) > 0:
                missing_hubs = [h for h in HUBS if not any(h.lower() in str(fh).lower() for fh in found_hubs)]
                if missing_hubs:
                    print(f"\n⚠️ Hubs not found in data ({len(missing_hubs)}):")
                    for hub in missing_hubs:
                        print(f"   ✗ {hub}")
            
            # Show summary of extracted columns
            print(f"\n📊 Final columns in report ({len(final_df.columns)}):")
            for i, col in enumerate(final_df.columns, 1):
                print(f"   {i}. {col}")
            
            return final_df
        else:
            print("\n⚠️ Warning: Could not identify hub column")
            print("   Returning all data without filtering")
            print("   Available columns:", list(df.columns))
            return df
        
    except Exception as e:
        print(f"❌ Error extracting data: {e}")
        raise


def generate_report(df, output_file):
    """Generate Excel report from DataFrame"""
    try:
        if df.empty:
            print("⚠️ Warning: No data to export")
            return
        
        print(f"\n📝 Generating report: {output_file}")
        
        # Create Excel writer with formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='South_COD_Data', index=False)
            
            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['South_COD_Data']
            
            # Format header row
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✅ Report generated successfully: {output_file}")
        print(f"📁 File location: {os.path.abspath(output_file)}")
        
    except Exception as e:
        print(f"❌ Error generating report: {e}")
        raise


def display_summary(df):
    """Display summary statistics of the data"""
    try:
        if df.empty:
            print("\n⚠️ No data to summarize")
            return
        
        print(f"\n{'='*60}")
        print("📊 DATA SUMMARY")
        print(f"{'='*60}")
        print(f"Total Rows (filtered by hubs): {len(df)}")
        print(f"Total Columns: {len(df.columns)}")
        print(f"Expected Hubs: {len(HUBS)}")
        
        # Find hub column for summary
        hub_column = find_hub_column(df)
        if hub_column:
            unique_hubs = df[hub_column].nunique()
            print(f"Unique Hubs Found: {unique_hubs}")
        
        # Display first few rows
        print(f"\n{'='*60}")
        print("📋 PREVIEW (First 5 rows)")
        print(f"{'='*60}")
        print(df.head().to_string())
        
        # Display statistics for the extracted columns
        extracted_cols = ['Total Collection', 'Total Deposit', 'Overall Gap']
        available_cols = [col for col in extracted_cols if col in df.columns]
        
        if available_cols:
            print(f"\n{'='*60}")
            print("📈 EXTRACTED COLUMNS SUMMARY")
            print(f"{'='*60}")
            
            # Convert to numeric and show statistics
            for col in available_cols:
                try:
                    numeric_series = pd.to_numeric(df[col], errors='coerce')
                    print(f"\n{col}:")
                    print(f"   Total: {numeric_series.sum():,.2f}")
                    print(f"   Average: {numeric_series.mean():,.2f}")
                    print(f"   Min: {numeric_series.min():,.2f}")
                    print(f"   Max: {numeric_series.max():,.2f}")
                    print(f"   Non-null values: {numeric_series.notna().sum()}/{len(df)}")
                except Exception as e:
                    print(f"\n{col}: Could not calculate statistics - {e}")
        
        # Display basic statistics for all numeric columns
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            print(f"\n{'='*60}")
            print("📈 ALL NUMERIC COLUMNS SUMMARY")
            print(f"{'='*60}")
            print(df[numeric_cols].describe().to_string())
        
    except Exception as e:
        print(f"⚠️ Error displaying summary: {e}")


def read_previous_actual_gap_values(worksheet, hub_column_name):
    """Read previous Actual Gap values from the worksheet before updating"""
    try:
        # Read existing data (skip header row, skip total row if exists)
        existing_data = worksheet.get_all_records()
        if not existing_data:
            return {}
        
        # Convert to DataFrame
        existing_df = pd.DataFrame(existing_data)
        
        # Find hub column
        hub_col = None
        for col in existing_df.columns:
            if str(col).strip().lower() in ['hub name', 'hub', 'hub_name']:
                hub_col = col
                break
        
        if not hub_col or CALCULATED_COLUMN not in existing_df.columns:
            return {}
        
        # Create dictionary mapping hub name to Actual Gap value
        previous_values = {}
        
        def safe_to_numeric(val):
            try:
                if pd.isna(val) or val == '' or val is None:
                    return 0
                if isinstance(val, str):
                    val = val.replace('₹', '').replace(',', '').replace(' ', '').strip()
                return pd.to_numeric(val, errors='coerce') or 0
            except:
                return 0
        
        for idx, row in existing_df.iterrows():
            hub_name = str(row[hub_col]).strip() if pd.notna(row[hub_col]) else ''
            if hub_name and hub_name.lower() != 'total':
                actual_gap_value = safe_to_numeric(row.get(CALCULATED_COLUMN, 0))
                if hub_name:
                    previous_values[hub_name] = actual_gap_value
        
        print(f"   📊 Read {len(previous_values)} previous Actual Gap values")
        return previous_values
    except Exception as e:
        print(f"   ⚠️ Error reading previous Actual Gap values: {e}")
        return {}


def compare_actual_gap_changes(new_df, previous_values, hub_column_name):
    """Compare new Actual Gap values with previous values and identify increases"""
    increases = []
    
    # Columns to exclude from email processing
    EXCLUDED_COLUMNS = ['Total Collection', 'Total Deposit']
    
    def safe_to_numeric(val):
        try:
            if pd.isna(val) or val == '' or val is None:
                return 0
            if isinstance(val, str):
                val = val.replace('₹', '').replace(',', '').replace(' ', '').strip()
            return pd.to_numeric(val, errors='coerce') or 0
        except:
            return 0
    
    if hub_column_name not in new_df.columns or CALCULATED_COLUMN not in new_df.columns:
        return increases
    
    # Ensure excluded columns are not part of the comparison (they shouldn't be, but explicit check)
    for col in EXCLUDED_COLUMNS:
        if col in new_df.columns:
            print(f"   ℹ️  Excluding '{col}' from email comparison (as requested)")
    
    # Create case-insensitive lookup for previous values
    previous_values_lower = {}
    for hub, value in previous_values.items():
        previous_values_lower[hub.lower().strip()] = value
        previous_values_lower[hub.strip()] = value  # Also keep original for exact match
    
    for idx, row in new_df.iterrows():
        hub_name = str(row[hub_column_name]).strip() if pd.notna(row[hub_column_name]) else ''
        if not hub_name or hub_name.lower() == 'total':
            continue
        
        # Only compare Actual Gap (exclude Total Collection and Total Deposit)
        new_value = round(safe_to_numeric(row.get(CALCULATED_COLUMN, 0)))
        
        # Try to find previous value (case-insensitive)
        previous_value = previous_values.get(hub_name, None)
        if previous_value is None:
            # Try case-insensitive lookup
            previous_value = previous_values_lower.get(hub_name.lower().strip(), 0)
        
        # Debug output for first few comparisons
        if idx < 3:
            print(f"      Comparing {hub_name}: Previous={previous_value}, New={new_value}")
        
        # Check if Actual Gap increased
        if new_value > previous_value:
            deviation = new_value - previous_value
            increases.append({
                'hub_name': hub_name,
                'previous_value': previous_value,
                'new_value': new_value,
                'deviation': deviation
            })
            print(f"      ⚠️ Increase detected: {hub_name} - ₹{previous_value:,} → ₹{new_value:,} (+₹{deviation:,})")
    
    return increases


def create_email_html_template(increases, df_data, hub_column_name, test_mode=False):
    """Create HTML email template with full table (same style as reservations_email_automation.py)"""
    today_date = datetime.now().strftime('%d-%b-%Y')
    time_str = datetime.now().strftime('%H:%M')
    
    # Create a set of hub names with increases for highlighting
    increased_hubs = {inc['hub_name'] for inc in increases} if increases else set()
    
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <meta http-equiv="Content-Type" content="text/html; charset=UTF-8">
    <style>
        body {{ 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            line-height: 1.6; 
            color: #2c3e50; 
            margin: 0;
            padding: 10px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
        }}
        .container {{
            max-width: 100%;
            width: 100%;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #FF6B35 0%, #F7931E 50%, #FFD23F 100%);
            color: white;
            padding: 12px 15px;
            text-align: center;
        }}
        .header h1 {{
            margin: 0;
            font-size: 16px;
            font-weight: bold;
        }}
        .header p {{
            margin: 5px 0 0 0;
            font-size: 11px;
        }}
        .content {{
            padding: 12px;
            overflow-x: auto;
        }}
        .table-wrapper {{
            overflow-x: auto;
            -webkit-overflow-scrolling: touch;
            width: 100%;
        }}
        .summary-section {{
            margin-bottom: 15px;
        }}
        .summary-title {{
            font-size: 15px;
            font-weight: bold;
            color: #FF6B35;
            margin-bottom: 10px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 11px;
            table-layout: auto;
            min-width: 100%;
        }}
        th {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 6px 4px;
            text-align: left;
            font-weight: 600;
            white-space: nowrap;
            font-size: 10px;
            overflow: hidden;
            text-overflow: ellipsis;
        }}
        th:first-child {{
            min-width: 100px;
            max-width: 130px;
        }}
        th:not(:first-child):not(:last-child):not(:nth-last-child(2)):not(:nth-last-child(3)) {{
            min-width: 45px;
            max-width: 55px;
        }}
        th:nth-last-child(3), th:nth-last-child(2), th:last-child {{
            min-width: 50px;
            max-width: 65px;
        }}
        td {{
            padding: 6px 4px;
            border-bottom: 1px solid #e0e0e0;
            background: white;
            font-size: 10px;
            white-space: nowrap;
        }}
        td:first-child {{
            min-width: 100px;
            max-width: 130px;
            word-wrap: break-word;
            white-space: normal;
        }}
        td:not(:first-child):not(:last-child):not(:nth-last-child(2)):not(:nth-last-child(3)) {{
            min-width: 45px;
            max-width: 55px;
            text-align: center;
        }}
        td:nth-last-child(3), td:nth-last-child(2), td:last-child {{
            min-width: 50px;
            max-width: 65px;
            text-align: center;
        }}
        @media only screen and (max-width: 600px) {{
            body {{
                padding: 5px;
            }}
            .container {{
                border-radius: 10px;
            }}
            .header {{
                padding: 10px;
            }}
            .header h1 {{
                font-size: 12px;
            }}
            .header p {{
                font-size: 9px;
            }}
            .content {{
                padding: 3px;
            }}
            .summary-title {{
                font-size: 13px;
            }}
            table {{
                font-size: 9px;
                min-width: 100%;
            }}
            th, td {{
                padding: 4px 3px;
                font-size: 9px;
            }}
            th:first-child, td:first-child {{
                min-width: 80px;
                max-width: 100px;
                font-size: 9px;
            }}
            th:not(:first-child):not(:last-child), td:not(:first-child):not(:last-child) {{
                min-width: 45px;
                max-width: 55px;
                font-size: 9px;
            }}
            th:last-child, td:last-child {{
                min-width: 50px;
                max-width: 60px;
                font-size: 9px;
            }}
            th:nth-last-child(2), td:nth-last-child(2) {{
                min-width: 50px;
                max-width: 60px;
                font-size: 9px;
            }}
            th:nth-last-child(3), td:nth-last-child(3) {{
                min-width: 50px;
                max-width: 60px;
                font-size: 9px;
            }}
        }}
        tr:nth-child(even) td {{
            background: #f8f9fa;
        }}
        tr.increase-row td {{
            background: #ffebee !important;
            border-left: 3px solid #dc3545;
        }}
        tr.total-row td {{
            background: #e3f2fd !important;
            font-weight: bold;
            border-top: 3px solid #2196f3;
        }}
        tr.total-row td:first-child {{
            background: linear-gradient(135deg, #2196f3 0%, #1976d2 100%) !important;
            color: white;
        }}
        .no-increases {{
            background: #d4edda;
            color: #155724;
            padding: 12px;
            border-radius: 5px;
            border: 1px solid #c3e6cb;
            font-size: 13px;
        }}
        .footer {{
            background: #f5f5f5;
            padding: 12px 15px;
            text-align: center;
            font-size: 11px;
            color: #666;
        }}
        .link-button {{
            display: inline-block;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 10px 20px;
            text-decoration: none;
            border-radius: 5px;
            margin-top: 15px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 South COD Monitor - Actual Gap Alert</h1>
            <p>{today_date} at {time_str}</p>
        </div>
        <div class="content">
"""
    
    # Add summary message
    if increases and len(increases) > 0:
        # Create list of hubs with increases
        increases_list = ""
        for inc in increases:
            increases_list += f"<li><strong>{inc['hub_name']}</strong>: ₹{inc['previous_value']:,} → ₹{inc['new_value']:,} <span style='color: #dc3545; font-weight: bold;'>(+₹{inc['deviation']:,})</span></li>"
        
        html += f"""
            <div class="summary-section">
                <div class="summary-title" style="background: #ffebee; padding: 10px; border-left: 4px solid #dc3545; margin-bottom: 15px;">
                    ⚠️ ACTUAL GAP INCREASES DETECTED - {len(increases)} Hub(s) Affected
                </div>
                <p style="font-weight: bold; color: #dc3545;">The following hub(s) have shown an increase in Actual Gap (highlighted in red in the table below):</p>
                <ul style="background: #fff3cd; padding: 10px 10px 10px 30px; border-radius: 5px; margin: 10px 0;">
                    {increases_list}
                </ul>
            </div>
"""
    else:
        html += """
            <div class="summary-section">
                <div class="no-increases">
                    <strong>✅ No increases detected</strong><br>
                    All hubs are stable or showing improvement in Actual Gap.
                </div>
            </div>
"""
    
    # Create full data table
    if df_data is not None and not df_data.empty and hub_column_name in df_data.columns:
        html += """
            <div class="summary-section">
                <div class="summary-title">📊 Complete COD Monitor Data</div>
                <div class="table-wrapper">
                <table style="width: 100%; font-size: 11px; border-collapse: collapse; table-layout: auto;">
                    <tr>
"""
        # Add table headers (excluding Total Collection and Total Deposit)
        columns_to_show = ['Hub Name', 'Overall Gap', 
                          'Van Adhoc', 'Legal Issue', 'Old Balance', 'Actual Gap']
        # Add latest date columns if they exist
        latest_date_cols = [col for col in df_data.columns if 'Colc ' in str(col) or ('Gap ' in str(col) and col != 'Overall Gap')]
        columns_to_show.extend(latest_date_cols)
        
        for col in columns_to_show:
            if col in df_data.columns or col == 'Hub Name':
                html += f'                        <th style="font-size: 10px; padding: 6px 4px; white-space: nowrap;">{col}</th>\n'
        
        html += """                    </tr>
"""
        
        # Calculate and add Total row as first row
        def safe_to_numeric(val):
            try:
                if pd.isna(val) or val == '' or val is None:
                    return 0
                if isinstance(val, str):
                    val = val.replace(',', '').replace('₹', '').replace(' ', '').strip()
                return float(val) if val else 0
            except:
                return 0
        
        # Add Total row with special styling
        html += '                    <tr class="total-row">\n'
        
        for col in columns_to_show:
            if col == 'Hub Name':
                html += '                        <td style="font-weight: bold; background: linear-gradient(135deg, #2196f3 0%, #1976d2 100%); color: white; font-size: 10px; padding: 6px 4px;">Total</td>\n'
            elif col in df_data.columns:
                # Calculate sum for this column
                total_value = 0
                for idx, row in df_data.iterrows():
                    hub_name = str(row[hub_column_name]).strip() if pd.notna(row[hub_column_name]) else ''
                    if hub_name and hub_name.lower() != 'total':
                        value = row.get(col, '')
                        total_value += safe_to_numeric(value)
                
                # Format the total
                if total_value == int(total_value):
                    formatted_value = f'₹{int(total_value):,}'
                else:
                    formatted_value = f'₹{total_value:,.2f}'
                html += f'                        <td style="text-align: right; font-weight: bold; background: #e3f2fd; font-size: 10px; padding: 6px 4px;">{formatted_value}</td>\n'
            else:
                html += '                        <td style="text-align: right; font-weight: bold; background: #e3f2fd; font-size: 10px; padding: 6px 4px;">-</td>\n'
        
        html += '                    </tr>\n'
        
        # Add data rows (excluding total row)
        for idx, row in df_data.iterrows():
            hub_name = str(row[hub_column_name]).strip() if pd.notna(row[hub_column_name]) else ''
            if not hub_name or hub_name.lower() == 'total':
                continue
            
            # Check if this hub has an increase
            has_increase = hub_name in increased_hubs
            row_class = 'increase-row' if has_increase else ''
            
            html += f'                    <tr class="{row_class}">\n'
            
            for col in columns_to_show:
                if col == 'Hub Name':
                    html += f'                        <td style="font-size: 10px; padding: 6px 4px; max-width: 130px; word-wrap: break-word;">{hub_name}</td>\n'
                elif col in df_data.columns:
                    value = row.get(col, '')
                    if pd.isna(value) or value == '':
                        html += '                        <td style="text-align: right; font-size: 10px; padding: 6px 4px;">-</td>\n'
                    else:
                        try:
                            # Format as number with currency
                            num_value = float(str(value).replace(',', '').replace('₹', '').strip())
                            if num_value == int(num_value):
                                formatted_value = f'₹{int(num_value):,}'
                            else:
                                formatted_value = f'₹{num_value:,.2f}'
                            html += f'                        <td style="text-align: right; font-size: 10px; padding: 6px 4px;">{formatted_value}</td>\n'
                        except:
                            html += f'                        <td style="text-align: right; font-size: 10px; padding: 6px 4px;">{value}</td>\n'
                else:
                    html += '                        <td style="text-align: right; font-size: 10px; padding: 6px 4px;">-</td>\n'
            
            html += '                    </tr>\n'
        
        html += """                </table>
                </div>
            </div>
"""
    
    html += """        </div>
        <div class="footer">
            <p>This is an automated email from the South COD Monitor system. Please do not reply.</p>
            <p>Generated at: """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """</p>"""
    
    # Add test mode message only if test_mode is True
    if test_mode:
        html += """            <p style='color: #dc3545; font-weight: bold;'>⚠️ TEST MODE - This is a test email. Recipients have been muted.</p>"""
    
    html += """        </div>
    </div>
</body>
</html>"""
    
    return html


def send_email_with_summary(increases, df_data, hub_column_name, spreadsheet_url):
    """Send email with summary of Actual Gap increases (using same email config as reservations_email_automation.py)"""
    try:
        if not EMAIL_ENABLED or not EMAIL_CONFIG['sender_password']:
            print(f"\n📧 Email sending is disabled or not configured")
            return
        
        print(f"\n📧 Preparing email with summary...")
        
        # Get email recipients dynamically (from G-Form_COD_Status.py)
        to_recipients, cc_recipients, bcc_recipients = get_email_recipients()
        
        # Check test mode
        if TEST_MODE:
            print(f"   🧪 TEST MODE ENABLED - Recipients muted, sending to test email only")
            actual_to_recipients = [TEST_EMAIL]
            actual_cc_list = []
            actual_bcc_list = []
            print(f"   📧 Test recipient: {TEST_EMAIL}")
            print(f"   📧 Original recipients (muted): TO={len(to_recipients)} recipients, BCC={len(bcc_recipients)} recipients")
        else:
            # TO: All recipients (hubs + CLMs + Lokesh + Bharath + Maligai Rasmeen)
            actual_to_recipients = to_recipients
            # CC: Empty (no CC recipients)
            actual_cc_list = []
            # BCC: Rakib only
            actual_bcc_list = bcc_recipients
            print(f"   📧 Production mode - Sending to actual recipients")
            print(f"   📧 To: {len(actual_to_recipients)} recipients")
            print(f"   📧 CC: {len(actual_cc_list)} recipients")
            print(f"   📧 BCC: {len(actual_bcc_list)} recipient(s)")
        
        today_date = datetime.now().strftime('%d-%b')
        current_time = datetime.now().strftime('%H:%M')
        
        # Create HTML email content with full table
        html_content = create_email_html_template(increases, df_data, hub_column_name, test_mode=TEST_MODE)
        
        # Create plain text version for email clients that don't support HTML
        plain_text = f"""South COD Monitor - Actual Gap Alert
{today_date} - {current_time}

"""
        if increases:
            plain_text += f"ACTUAL GAP INCREASES DETECTED\n\n"
            plain_text += f"The following {len(increases)} hub(s) have shown an increase in Actual Gap:\n\n"
            for inc in increases:
                plain_text += f"{inc['hub_name']}: ₹{inc['previous_value']:,} -> ₹{inc['new_value']:,} (+₹{inc['deviation']:,})\n"
        else:
            plain_text += "No increases detected. All hubs are stable or showing improvement.\n"
        
        plain_text += f"\nView full report: {spreadsheet_url}\n"
        
        # Create email message with proper MIME structure for HTML preservation
        msg = MIMEMultipart('alternative')
        msg['From'] = EMAIL_CONFIG['sender_email']
        msg['To'] = ', '.join(actual_to_recipients)
        if actual_cc_list:
            msg['Cc'] = ', '.join(actual_cc_list)
        
        subject_prefix = "[TEST MODE] " if TEST_MODE else ""
        msg['Subject'] = f'{subject_prefix}South - COD Monitor - {today_date} - {current_time}'
        
        # Add headers to encourage HTML format preservation
        msg['X-Mailer'] = 'Python South COD Monitor'
        msg['MIME-Version'] = '1.0'
        
        # Attach plain text first (lower priority)
        part1 = MIMEText(plain_text, 'plain', 'utf-8')
        msg.attach(part1)
        
        # Attach HTML (higher priority - will be preferred by email clients)
        part2 = MIMEText(html_content, 'html', 'utf-8')
        part2.add_header('Content-Type', 'text/html; charset=utf-8')
        msg.attach(part2)
        
        # All recipients (To + CC + BCC for sending)
        all_recipients = actual_to_recipients + actual_cc_list + actual_bcc_list
        
        # Send email with retry logic (same as reservations_email_automation.py)
        max_retries = 3
        timeout = 60
        last_error = None
        
        for attempt in range(1, max_retries + 1):
            try:
                mode_indicator = "[TEST MODE] " if TEST_MODE else ""
                print(f"   📤 {mode_indicator}Attempt {attempt}/{max_retries}: Sending email...")
                print(f"      To: {len(actual_to_recipients)} recipient(s)")
                if actual_cc_list:
                    print(f"      CC: {len(actual_cc_list)} recipients")
                if actual_bcc_list:
                    print(f"      BCC: {len(actual_bcc_list)} recipient(s)")
                if TEST_MODE:
                    print(f"      ⚠️  Original recipients muted (not receiving email)")
                
                server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'], timeout=timeout)
                server.timeout = timeout
                server.starttls()
                server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
                
                text = msg.as_string()
                server.sendmail(EMAIL_CONFIG['sender_email'], all_recipients, text)
                server.quit()
                
                print(f"   ✅ Email sent successfully")
                if increases:
                    print(f"   📊 Summary: {len(increases)} hub(s) with increased Actual Gap")
                    for inc in increases:
                        print(f"      • {inc['hub_name']}: ₹{inc['previous_value']:,} → ₹{inc['new_value']:,} (+₹{inc['deviation']:,})")
                return  # Success, exit function
                
            except (smtplib.SMTPConnectError, smtplib.SMTPException, ConnectionError, TimeoutError, OSError) as e:
                last_error = e
                print(f"   ⚠️ Attempt {attempt}/{max_retries} failed: {e}")
                if attempt < max_retries:
                    wait_time = attempt * 5
                    print(f"   ⏳ Waiting {wait_time} seconds before retry...")
                    time.sleep(wait_time)
                else:
                    print(f"   ❌ Failed to send email after {max_retries} attempts")
                    
            except smtplib.SMTPAuthenticationError as e:
                print(f"   ❌ Authentication failed: {e}")
                print(f"   Please check your Gmail App Password")
                raise
                
            except Exception as e:
                last_error = e
                print(f"   ❌ Unexpected error: {e}")
                if attempt < max_retries:
                    wait_time = attempt * 5
                    time.sleep(wait_time)
                else:
                    raise
        
        if last_error:
            raise last_error
        
    except Exception as e:
        print(f"   ❌ Error sending email: {e}")
        import traceback
        traceback.print_exc()


def upload_to_google_sheets(df, client):
    """Upload DataFrame to Google Sheets, preserving manually editable columns"""
    # Initialize variables for email summary
    increases_summary = []
    previous_actual_gap_values = {}
    
    try:
        if df.empty:
            print("⚠️ Warning: No data to upload to Google Sheets")
            return
        
        print(f"\n{'='*60}")
        print("📤 UPLOADING TO GOOGLE SHEETS")
        print(f"{'='*60}")
        print(f"📋 Spreadsheet ID: {OUTPUT_SPREADSHEET_ID}")
        print(f"📄 Worksheet Name: {OUTPUT_WORKSHEET_NAME}")
        
        # Open the output spreadsheet
        spreadsheet = client.open_by_key(OUTPUT_SPREADSHEET_ID)
        print(f"✅ Opened spreadsheet: {spreadsheet.title}")
        
        # Get or create the worksheet
        try:
            worksheet = spreadsheet.worksheet(OUTPUT_WORKSHEET_NAME)
            
            # Before clearing, preserve the manually editable columns
            print(f"\n🔒 Preserving manually editable columns: {', '.join(PRESERVE_COLUMNS)}")
            preserved_data = {}
            
            try:
                # Read existing data directly using get_all_values() to avoid duplicate header issues
                all_sheet_values = worksheet.get_all_values()
                print(f"   📊 Worksheet has {len(all_sheet_values)} rows")
                
                if len(all_sheet_values) < 2:
                    print(f"   ⚠️  Worksheet appears empty or has no data rows (only {len(all_sheet_values)} row(s))")
                    print(f"   ℹ️  This might be the first run - no data to preserve")
                    existing_df = pd.DataFrame()
                    hub_col_existing = None
                else:
                    print(f"   ✅ Worksheet has data - proceeding to preserve values")
                    
                    # Parse data manually: Row 0 = Total, Row 1 = Header, Row 2+ = Data
                    header_row = all_sheet_values[1] if len(all_sheet_values) > 1 else []
                    data_rows = all_sheet_values[2:] if len(all_sheet_values) > 2 else []
                    
                    print(f"   📋 Header row: {header_row[:5]}...")  # Show first 5 headers
                    print(f"   📊 Data rows: {len(data_rows)} rows")
                    
                    # Create DataFrame manually to avoid duplicate header issues
                    if header_row and data_rows:
                        # Make headers unique if there are duplicates
                        unique_headers = []
                        header_count = {}
                        for h in header_row:
                            h_str = str(h).strip()
                            if h_str in header_count:
                                header_count[h_str] += 1
                                unique_headers.append(f"{h_str}_{header_count[h_str]}")
                            else:
                                header_count[h_str] = 0
                                unique_headers.append(h_str)
                        
                        # Create DataFrame with unique headers
                        existing_df = pd.DataFrame(data_rows, columns=unique_headers)
                        print(f"   📊 DataFrame created with {len(existing_df)} rows and {len(existing_df.columns)} columns")
                        
                        # Find hub column in existing data (check original header names)
                        hub_col_existing = None
                        for idx, col in enumerate(existing_df.columns):
                            original_header = header_row[idx] if idx < len(header_row) else col
                            if str(original_header).strip().lower() in ['hub name', 'hub', 'hub_name']:
                                hub_col_existing = col
                                print(f"   ✅ Found Hub column: '{col}' (original: '{original_header}')")
                                break
                        
                        # Read previous Actual Gap values for comparison (before clearing)
                        if hub_col_existing:
                            previous_actual_gap_values = read_previous_actual_gap_values(worksheet, hub_col_existing)
                    else:
                        existing_df = pd.DataFrame()
                        hub_col_existing = None
                    
                    if hub_col_existing and len(all_sheet_values) >= 2:
                        # Extract preserved columns data - use the already-read all_sheet_values
                        # Row 0 = Total, Row 1 = Header, Row 2+ = Data
                        header_row = all_sheet_values[1] if len(all_sheet_values) > 1 else []
                        data_rows = all_sheet_values[2:] if len(all_sheet_values) > 2 else []
                        
                        print(f"   📋 Header row: {header_row[:10]}...")  # Show first 10 headers
                        print(f"   📊 Data rows: {len(data_rows)} rows found")
                        
                        # Find column indices in the original header row
                        hub_col_idx = None
                        preserve_col_indices = {}
                        
                        print(f"   🔍 Searching for columns in header row...")
                        for idx, header in enumerate(header_row):
                            header_str = str(header).strip()
                            header_lower = header_str.lower()
                            
                            # Check for hub column
                            if header_lower in ['hub name', 'hub', 'hub_name']:
                                hub_col_idx = idx
                                print(f"      ✅ Found Hub column at index {idx}: '{header_str}'")
                            
                            # Check for preserve columns (exact match)
                            for preserve_col in PRESERVE_COLUMNS:
                                if header_str == preserve_col:
                                    preserve_col_indices[preserve_col] = idx
                                    print(f"      ✅ Found '{preserve_col}' column at index {idx}")
                        
                        # Debug: Show what columns we're looking for vs what we found
                        print(f"   📋 Looking for preserve columns: {PRESERVE_COLUMNS}")
                        print(f"   📋 Found preserve columns: {list(preserve_col_indices.keys())}")
                        if not preserve_col_indices:
                            print(f"   ⚠️  No preserve columns found! Available headers: {header_row}")
                        
                        if hub_col_idx is None:
                            print(f"   ⚠️  Hub column not found! Available headers: {header_row[:10]}")
                        
                        # Read preserved values directly from sheet rows
                        for preserve_col in PRESERVE_COLUMNS:
                            if preserve_col in preserve_col_indices and hub_col_idx is not None:
                                col_idx = preserve_col_indices[preserve_col]
                                preserved_data[preserve_col] = {}
                                
                                print(f"   📖 Reading '{preserve_col}' from column index {col_idx}...")
                                rows_read = 0
                                
                                for row_idx, row in enumerate(data_rows):
                                    if hub_col_idx < len(row):
                                        hub_name = str(row[hub_col_idx]).strip() if hub_col_idx < len(row) and row[hub_col_idx] else ''
                                        # Skip total row and empty hub names
                                        if hub_name and hub_name.lower() != 'total':
                                            # Get raw value from sheet
                                            if col_idx < len(row):
                                                raw_value = row[col_idx]
                                                
                                                # Preserve ALL values including 0, empty strings, etc.
                                                # This ensures manually edited values are always preserved
                                                if raw_value == '' or raw_value is None:
                                                    # Empty cell - preserve as 0 (user may have cleared it)
                                                    preserved_value = 0
                                                else:
                                                    try:
                                                        # Clean and convert to number
                                                        clean_value = str(raw_value).replace(',', '').replace('₹', '').replace(' ', '').strip()
                                                        if clean_value == '':
                                                            preserved_value = 0
                                                        else:
                                                            # Convert to float, then to int if it's a whole number
                                                            num_value = float(clean_value)
                                                            if num_value == int(num_value):
                                                                preserved_value = int(num_value)
                                                            else:
                                                                preserved_value = num_value
                                                    except (ValueError, AttributeError):
                                                        # If conversion fails, preserve as string (shouldn't happen for numeric columns)
                                                        preserved_value = str(raw_value).strip()
                                                
                                                # ALWAYS preserve the value (including 0) - this is a manually editable column
                                                preserved_data[preserve_col][hub_name] = preserved_value
                                                rows_read += 1
                                                
                                                # Debug first few values
                                                if rows_read <= 3:
                                                    print(f"      Row {row_idx+3}: {hub_name} = {raw_value} -> {preserved_value}")
                                            else:
                                                print(f"      ⚠️  Row {row_idx+3}: Column index {col_idx} out of range (row length: {len(row)})")
                                
                                print(f"   📊 Read {rows_read} values for '{preserve_col}'")
                                
                                # Count preserved values by type
                                total_count = len(preserved_data[preserve_col])
                                zero_count = sum(1 for v in preserved_data[preserve_col].values() if v == 0 or v == '0')
                                non_zero_count = total_count - zero_count
                                
                                print(f"   ✅ Preserved '{preserve_col}': {total_count} values ({non_zero_count} non-zero, {zero_count} zeros)")
                                
                                # Debug: Show all preserved values (not just samples) to verify they're captured
                                if preserved_data[preserve_col]:
                                    print(f"      📋 All preserved values for '{preserve_col}':")
                                    for hub, val in sorted(preserved_data[preserve_col].items()):
                                        print(f"         • {hub}: {val}")
                            elif preserve_col in existing_df.columns and not existing_df.empty:
                                # Fallback: use DataFrame if column index not found
                                print(f"   ⚠️  Column '{preserve_col}' not found via direct reading, using DataFrame fallback...")
                                preserved_data[preserve_col] = {}
                                for idx, row in existing_df.iterrows():
                                    hub_name = str(row[hub_col_existing]).strip() if pd.notna(row[hub_col_existing]) else ''
                                    if hub_name and hub_name.lower() != 'total':
                                        preserve_value = row.get(preserve_col)
                                        # Preserve the value as-is, including 0
                                        if pd.isna(preserve_value) or preserve_value is None:
                                            preserve_value = 0
                                        else:
                                            try:
                                                # Convert to number if possible
                                                if isinstance(preserve_value, str):
                                                    clean = preserve_value.replace(',', '').replace('₹', '').strip()
                                                    preserve_value = float(clean) if clean else 0
                                                preserve_value = int(preserve_value) if preserve_value == int(preserve_value) else preserve_value
                                            except:
                                                preserve_value = preserve_value
                                        
                                        preserved_data[preserve_col][hub_name] = preserve_value
                                
                                total_count = len(preserved_data[preserve_col])
                                zero_count = sum(1 for v in preserved_data[preserve_col].values() if v == 0 or v == '0')
                                non_zero_count = total_count - zero_count
                                print(f"   ✅ Preserved '{preserve_col}': {total_count} values ({non_zero_count} non-zero, {zero_count} zeros) using DataFrame fallback")
                                
                                # Show all preserved values
                                if preserved_data[preserve_col]:
                                    print(f"      📋 All preserved values for '{preserve_col}':")
                                    for hub, val in sorted(preserved_data[preserve_col].items()):
                                        print(f"         • {hub}: {val}")
                            else:
                                print(f"   ⚠️ Column '{preserve_col}' not found in existing data")
                                if not existing_df.empty:
                                    print(f"      Available columns: {list(existing_df.columns)}")
                    
                    # Summary of all preserved data
                    if preserved_data:
                        total_preserved = sum(len(v) for v in preserved_data.values())
                        print(f"\n   📊 Preservation Summary: {total_preserved} total values preserved across {len(preserved_data)} columns")
                        for col, values in preserved_data.items():
                            non_zero = sum(1 for v in values.values() if v != 0 and v != '0')
                            zeros = len(values) - non_zero
                            print(f"      • {col}: {len(values)} values ({non_zero} non-zero, {zeros} zeros)")
                    else:
                        print(f"   ⚠️ Could not find hub column in existing data to preserve columns")
            except Exception as e:
                print(f"   ⚠️ Could not read existing data to preserve columns: {e}")
                print(f"   Continuing with upload (new worksheet or empty sheet)")
            
            # Clear existing data completely
            worksheet.clear()
            print(f"✅ Cleared existing data from worksheet '{OUTPUT_WORKSHEET_NAME}'")
            time.sleep(1)  # Small delay to ensure clear operation completes
        except gspread.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(title=OUTPUT_WORKSHEET_NAME, rows=1000, cols=20)
            print(f"✅ Created new worksheet '{OUTPUT_WORKSHEET_NAME}'")
        
        # Convert all columns to standard Python types to avoid JSON serialization errors
        print("🔄 Preparing data for upload...")
        def convert_to_serializable(obj):
            if pd.isna(obj) or obj is None:
                return None
            elif isinstance(obj, (np.integer, np.int64, np.int32)):
                return int(obj)
            elif isinstance(obj, (np.floating, np.float64, np.float32)):
                return float(obj)
            elif isinstance(obj, np.bool_):
                return bool(obj)
            else:
                return obj
        
        # Apply conversion to all DataFrame values
        df_upload = df.map(convert_to_serializable).copy()
        
        # Find hub column in the upload data
        hub_col_upload = None
        for col in df_upload.columns:
            if str(col).strip().lower() in ['hub name', 'hub', 'hub_name']:
                hub_col_upload = col
                break
        
        # Add preserved columns to the upload DataFrame
        if hub_col_upload and preserved_data:
            print(f"\n📝 Merging preserved columns into upload data...")
            for preserve_col in PRESERVE_COLUMNS:
                # Initialize column with None (will be filled with preserved values or 0)
                df_upload[preserve_col] = None
                
                # Map preserved values based on hub name (case-insensitive matching)
                if preserve_col in preserved_data and preserved_data[preserve_col]:
                    # Create case-insensitive lookup dictionary
                    preserved_lookup = {}
                    for hub_name, value in preserved_data[preserve_col].items():
                        # Store with original hub name (case-sensitive)
                        preserved_lookup[hub_name] = value
                        # Also store with lowercase for case-insensitive matching
                        preserved_lookup[hub_name.lower()] = value
                    
                    def get_preserved_value(hub_name):
                        if not hub_name or pd.isna(hub_name):
                            return 0  # Default to 0 if no hub name
                        hub_str = str(hub_name).strip()
                        # Try exact match first (case-sensitive)
                        if hub_str in preserved_lookup:
                            val = preserved_lookup[hub_str]
                            # Return the preserved value as-is (including 0)
                            # Only convert None/NaN to 0, but preserve 0 as 0
                            if val is None or (isinstance(val, float) and pd.isna(val)):
                                return 0
                            # Preserve 0, empty string, or any other value
                            return val
                        # Try case-insensitive match
                        hub_lower = hub_str.lower()
                        if hub_lower in preserved_lookup:
                            val = preserved_lookup[hub_lower]
                            if val is None or (isinstance(val, float) and pd.isna(val)):
                                return 0
                            return val
                        # If no match found, return 0 (new hub or no previous data)
                        return 0
                    
                    df_upload[preserve_col] = df_upload[hub_col_upload].apply(get_preserved_value)
                    
                    # Count and report merged values
                    total_hubs = len(df_upload[df_upload[hub_col_upload].notna() & (df_upload[hub_col_upload].str.strip().str.lower() != 'total')])
                    non_zero_count = (df_upload[preserve_col] != 0).sum()
                    zero_count = (df_upload[preserve_col] == 0).sum()
                    
                    print(f"   ✅ Merged '{preserve_col}': {total_hubs} hubs processed ({non_zero_count} non-zero, {zero_count} zeros)")
                    
                    # Show merged values for first few hubs to verify preservation
                    print(f"      📋 Sample merged values (first 5 hubs):")
                    sample_count = 0
                    for idx, row in df_upload.iterrows():
                        if sample_count >= 5:
                            break
                        hub_name = str(row[hub_col_upload]).strip() if pd.notna(row[hub_col_upload]) else ''
                        if hub_name and hub_name.lower() != 'total':
                            value = row[preserve_col]
                            print(f"         • {hub_name}: {value}")
                            sample_count += 1
                else:
                    # No preserved data - initialize with 0
                    df_upload[preserve_col] = 0
                    print(f"   ⚠️ No preserved data for '{preserve_col}' - initialized with 0")
        
        # Calculate "Actual Gap" = Overall Gap - (Van Adhoc + Legal Issue + Old Balance + Gap Date)
        print(f"\n🧮 Calculating '{CALCULATED_COLUMN}'...")
        try:
            # Ensure all required columns exist
            if 'Overall Gap' not in df_upload.columns:
                print(f"   ⚠️ 'Overall Gap' column not found, cannot calculate '{CALCULATED_COLUMN}'")
            else:
                # Ensure preserved columns exist (fill None/NaN with 0 only for calculation, but keep original preserved values)
                for col in PRESERVE_COLUMNS:
                    if col not in df_upload.columns:
                        df_upload[col] = 0
                    else:
                        # Fill NaN/None with 0 only for calculation purposes, but preserve existing values
                        # This ensures calculation works even if some hubs don't have preserved values
                        df_upload[col] = df_upload[col].fillna(0)
                
                # Find the Gap date column (e.g., "Gap 11-Dec")
                gap_date_col = None
                for col in df_upload.columns:
                    if 'Gap ' in str(col) and col != 'Overall Gap':
                        gap_date_col = col
                        break
                
                if gap_date_col:
                    print(f"   📅 Found Gap date column: '{gap_date_col}'")
                else:
                    print(f"   ⚠️ Gap date column not found, will use 0 for calculation")
                
                # Convert to numeric, handling errors
                def safe_to_numeric(val):
                    try:
                        if pd.isna(val) or val == '' or val is None:
                            return 0
                        # Remove currency symbols and commas
                        if isinstance(val, str):
                            val = val.replace('₹', '').replace(',', '').replace(' ', '').strip()
                        return pd.to_numeric(val, errors='coerce') or 0
                    except:
                        return 0
                
                # Calculate Actual Gap = Overall Gap - (Van Adhoc + Legal Issue + Old Balance + Gap Date)
                overall_gap = df_upload['Overall Gap'].apply(safe_to_numeric)
                van_adhoc = df_upload['Van Adhoc'].apply(safe_to_numeric)
                legal_issue = df_upload['Legal Issue'].apply(safe_to_numeric)
                old_balance = df_upload['Old Balance'].apply(safe_to_numeric)
                
                # Include Gap date column in calculation
                if gap_date_col:
                    gap_date = df_upload[gap_date_col].apply(safe_to_numeric)
                    df_upload[CALCULATED_COLUMN] = overall_gap - (van_adhoc + legal_issue + old_balance + gap_date)
                    print(f"   📝 Formula: Overall Gap - (Van Adhoc + Legal Issue + Old Balance + {gap_date_col})")
                else:
                    df_upload[CALCULATED_COLUMN] = overall_gap - (van_adhoc + legal_issue + old_balance)
                    print(f"   📝 Formula: Overall Gap - (Van Adhoc + Legal Issue + Old Balance) [Gap Date column not found]")
                
                # Count non-zero values
                non_zero_count = (df_upload[CALCULATED_COLUMN] != 0).sum()
                print(f"   ✅ Calculated '{CALCULATED_COLUMN}': {non_zero_count} non-zero values")
                
        except Exception as e:
            print(f"   ⚠️ Error calculating '{CALCULATED_COLUMN}': {e}")
            # Add empty column if calculation fails
            if CALCULATED_COLUMN not in df_upload.columns:
                df_upload[CALCULATED_COLUMN] = None
        
        # Compare Actual Gap changes and prepare email summary
        increases_summary = []
        if hub_col_upload and CALCULATED_COLUMN in df_upload.columns:
            if previous_actual_gap_values:
                print(f"\n📊 Comparing Actual Gap changes...")
                print(f"   📋 Previous values found for {len(previous_actual_gap_values)} hubs")
                increases_summary = compare_actual_gap_changes(df_upload, previous_actual_gap_values, hub_col_upload)
                if increases_summary:
                    print(f"   ⚠️ Found {len(increases_summary)} hub(s) with increased Actual Gap:")
                    for inc in increases_summary:
                        print(f"      • {inc['hub_name']}: ₹{inc['previous_value']:,} → ₹{inc['new_value']:,} (+₹{inc['deviation']:,})")
                else:
                    print(f"   ✅ No increases in Actual Gap detected")
            else:
                print(f"\n📊 No previous Actual Gap values found (first run or empty sheet)")
                print(f"   ℹ️  Email will be sent without increase comparison")
        
        # Reorder columns to ensure correct order: Colc Date and Gap Date should be LAST
        print(f"\n📋 Reordering columns for final upload...")
        current_columns = list(df_upload.columns)
        
        # Identify latest date columns (Colc Date and Gap Date)
        latest_date_cols = [col for col in current_columns if 'Colc ' in str(col) or ('Gap ' in str(col) and col != 'Overall Gap')]
        # Filter to only those that match the pattern (not "Overall Gap")
        latest_date_cols = [col for col in latest_date_cols if col not in ['Overall Gap']]
        
        # Remove latest date columns from current position
        remaining_columns = [col for col in current_columns if col not in latest_date_cols]
        
        # Define the desired order:
        # 1. Hub Name (first)
        # 2. Total Collection, Total Deposit, Overall Gap
        # 3. Van Adhoc, Legal Issue, Old Balance (preserved columns)
        # 4. Actual Gap (calculated column)
        # 5. Colc Date, Gap Date (latest date columns - LAST)
        
        ordered_columns = []
        
        # Add Hub Name first if it exists
        hub_col = None
        for col in remaining_columns:
            if str(col).strip().lower() in ['hub name', 'hub', 'hub_name']:
                hub_col = col
                ordered_columns.append(col)
                break
        
        # Add standard columns (Total Collection, Total Deposit, Overall Gap)
        standard_cols = ['Total Collection', 'Total Deposit', 'Overall Gap']
        for col in standard_cols:
            if col in remaining_columns and col not in ordered_columns:
                ordered_columns.append(col)
        
        # Add preserved columns (Van Adhoc, Legal Issue, Old Balance)
        for col in PRESERVE_COLUMNS:
            if col in remaining_columns and col not in ordered_columns:
                ordered_columns.append(col)
        
        # Add calculated column (Actual Gap)
        if CALCULATED_COLUMN in remaining_columns and CALCULATED_COLUMN not in ordered_columns:
            ordered_columns.append(CALCULATED_COLUMN)
        
        # Add any remaining columns that haven't been added yet (except latest date columns)
        for col in remaining_columns:
            if col not in ordered_columns:
                ordered_columns.append(col)
        
        # Finally, add latest date columns at the end (LAST)
        ordered_columns.extend(latest_date_cols)
        
        # Reorder the DataFrame
        df_upload = df_upload[ordered_columns]
        print(f"   ✅ Columns reordered. Latest date columns placed at the end: {latest_date_cols}")
        print(f"   📋 Final column order: {list(df_upload.columns)}")
        
        # Sort by Actual Gap (descending - highest gap first)
        if CALCULATED_COLUMN in df_upload.columns:
            print(f"\n📊 Sorting data by '{CALCULATED_COLUMN}' (descending)...")
            # Helper function to convert to numeric for sorting
            def safe_to_numeric_for_sort(val):
                try:
                    if pd.isna(val) or val == '' or val is None:
                        return 0
                    if isinstance(val, str):
                        val = val.replace('₹', '').replace(',', '').replace(' ', '').strip()
                    return pd.to_numeric(val, errors='coerce') or 0
                except:
                    return 0
            
            # Convert Actual Gap to numeric for sorting
            df_upload['_sort_key'] = df_upload[CALCULATED_COLUMN].apply(safe_to_numeric_for_sort)
            # Sort by Actual Gap in descending order (highest gap first)
            df_upload = df_upload.sort_values('_sort_key', ascending=False)
            # Remove temporary sort key column
            df_upload = df_upload.drop(columns=['_sort_key'])
            print(f"   ✅ Data sorted by '{CALCULATED_COLUMN}' in descending order (highest to lowest)")
        else:
            print(f"\n⚠️ '{CALCULATED_COLUMN}' column not found, skipping sort")
        
        # Helper function to convert to numeric
        def safe_to_numeric(val):
            try:
                if pd.isna(val) or val == '' or val is None:
                    return 0
                if isinstance(val, str):
                    val = val.replace('₹', '').replace(',', '').replace(' ', '').strip()
                return pd.to_numeric(val, errors='coerce') or 0
            except:
                return 0
        
        # Round all numeric values in df_upload first
        print(f"\n🔢 Rounding all numeric values...")
        numeric_cols_list = ['Total Collection', 'Total Deposit', 'Overall Gap', CALCULATED_COLUMN] + PRESERVE_COLUMNS
        # Add latest date columns
        latest_date_cols_for_rounding = [col for col in df_upload.columns if 'Colc ' in str(col) or ('Gap ' in str(col) and col != 'Overall Gap')]
        numeric_cols_list.extend(latest_date_cols_for_rounding)
        
        for col_name in df_upload.columns:
            if col_name in numeric_cols_list:
                # Round numeric columns to nearest integer
                df_upload[col_name] = df_upload[col_name].apply(
                    lambda x: round(safe_to_numeric(x)) if pd.notna(x) and x != '' else x
                )
        
        print(f"   ✅ All numeric values rounded to whole numbers")
        
        # Create total row
        print(f"\n📊 Creating total row...")
        total_row = {}
        
        # Find hub name column (first column)
        hub_col = ordered_columns[0] if ordered_columns else None
        
        # Set "Total" in the hub name column
        for col in ordered_columns:
            if col == hub_col:
                total_row[col] = "Total"
            else:
                # For numeric columns, calculate sum and round
                try:
                    # Convert column to numeric and sum
                    numeric_values = df_upload[col].apply(safe_to_numeric)
                    total_row[col] = round(numeric_values.sum())
                except:
                    # For non-numeric columns, leave empty or use appropriate value
                    total_row[col] = ""
        
        # Create DataFrame with total row
        total_df = pd.DataFrame([total_row])
        print(f"   ✅ Total row created")
        
        # Upload: Total row first (row 1), then header + data (row 2 onwards)
        print(f"\n📤 Uploading data to Google Sheets...")
        # Upload total row first at row 1
        set_with_dataframe(worksheet, total_df, row=1, include_column_header=False)
        # Upload header + data starting from row 2
        set_with_dataframe(worksheet, df_upload, row=2, include_column_header=True)
        print("✅ Data uploaded successfully")
        
        # Apply formatting
        print("🎨 Applying formatting...")
        num_columns = len(df_upload.columns)
        num_data_rows = len(df_upload)
        
        # Helper function to convert column index to letter (A, B, ..., Z, AA, AB, etc.)
        def get_column_letter(col_idx):
            """Convert 0-based column index to Excel column letter"""
            result = ""
            col_idx += 1  # Convert to 1-based
            while col_idx > 0:
                col_idx -= 1
                result = string.ascii_uppercase[col_idx % 26] + result
                col_idx //= 26
            return result
        
        # Calculate last column letter
        last_col_letter = get_column_letter(num_columns - 1)
        
        # Format total row (row 1) with bold text and yellow background
        worksheet.format(f'A1:{last_col_letter}1', {
            'textFormat': {'bold': True},
            'backgroundColor': {'red': 1.0, 'green': 1.0, 'blue': 0.0}  # Yellow background
        })
        
        # Set left alignment for Hub Name column in total row (row 1)
        if num_columns >= 1:
            worksheet.format('A1', {
                'textFormat': {'bold': True},
                'backgroundColor': {'red': 1.0, 'green': 1.0, 'blue': 0.0},  # Yellow background
                'horizontalAlignment': 'LEFT'
            })
        
        # Format header row (row 2) - green background, bold, center aligned
        worksheet.format(f'A2:{last_col_letter}2', {
            'backgroundColor': {'red': 0.2, 'green': 0.8, 'blue': 0.2},
            'textFormat': {'bold': True},
            'horizontalAlignment': 'CENTER'
        })
        
        # Set left alignment for Hub Name column in header (row 2)
        if num_columns >= 1:
            worksheet.format('A2', {
                'backgroundColor': {'red': 0.2, 'green': 0.8, 'blue': 0.2},
                'textFormat': {'bold': True},
                'horizontalAlignment': 'LEFT'
            })
        
        # Format numeric columns as currency (Total Collection, Total Deposit, Overall Gap, Actual Gap, preserved columns, and latest date columns)
        numeric_cols = ['Total Collection', 'Total Deposit', 'Overall Gap', CALCULATED_COLUMN] + PRESERVE_COLUMNS
        # Add latest date columns (Colc Date and Gap Date)
        latest_date_cols = [col for col in df_upload.columns if 'Colc ' in str(col) or ('Gap ' in str(col) and col != 'Overall Gap')]
        numeric_cols.extend(latest_date_cols)
        
        for col_name in numeric_cols:
            if col_name in df_upload.columns:
                col_idx = list(df_upload.columns).index(col_name)
                col_letter = get_column_letter(col_idx)
                
                # Format as currency (₹) without decimals and right-aligned
                # Row 1 is total row, Row 2 is header, Row 3 onwards is data
                # Format total row (row 1) with yellow background
                worksheet.format(f'{col_letter}1', {
                    'numberFormat': {'type': 'CURRENCY', 'pattern': '₹#,##0'},
                    'horizontalAlignment': 'RIGHT',
                    'textFormat': {'bold': True},
                    'backgroundColor': {'red': 1.0, 'green': 1.0, 'blue': 0.0}  # Yellow background
                })
                
                # Format data rows (row 3 onwards)
                if num_data_rows > 0:
                    worksheet.format(f'{col_letter}3:{col_letter}{num_data_rows + 2}', {
                        'numberFormat': {'type': 'CURRENCY', 'pattern': '₹#,##0'},
                        'horizontalAlignment': 'RIGHT'
                    })
        
        # Left align Hub Name column in data rows (row 3 onwards)
        if num_columns >= 1 and num_data_rows > 0:
            worksheet.format(f'A3:A{num_data_rows + 2}', {
                'horizontalAlignment': 'LEFT'
            })
        
        print("✅ Formatting applied successfully")
        spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{OUTPUT_SPREADSHEET_ID}/edit#gid={worksheet.id}"
        print(f"\n🔗 Google Sheet URL: {spreadsheet_url}")
        
        # Send email with summary of Actual Gap increases
        # Use df_upload (before total row is added) for email table
        # Always send email (even if no increases) to show the full report
        if hub_col_upload:
            print(f"\n📧 Preparing to send email...")
            if increases_summary:
                print(f"   ⚠️ Alert: {len(increases_summary)} hub(s) with increased Actual Gap will be highlighted")
            else:
                print(f"   ℹ️  No increases detected - email will show full report without alerts")
            send_email_with_summary(increases_summary, df_upload, hub_col_upload, spreadsheet_url)
        else:
            print(f"\n⚠️ Cannot send email: Hub column not found")
        
    except Exception as e:
        print(f"❌ Error uploading to Google Sheets: {e}")
        import traceback
        traceback.print_exc()
        raise


def main():
    """Main function to run the South COD Monitor"""
    print("="*60)
    print("🚀 SOUTH COD MONITOR")
    print("="*60)
    print(f"📅 Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    try:
        # Setup Google Sheets connection
        client = setup_google_sheets()
        
        # Get the worksheet by name
        worksheet = get_worksheet_by_name(client, SPREADSHEET_ID, SOURCE_WORKSHEET_NAME)
        
        if not worksheet:
            print("❌ Could not access worksheet. Exiting...")
            sys.exit(1)
        
        # Extract data
        df = extract_sheet_data(worksheet)
        
        if df.empty:
            print("❌ No data extracted. Exiting...")
            sys.exit(1)
        
        # Display summary
        display_summary(df)
        
        # Generate Excel report
        generate_report(df, OUTPUT_FILE)
        
        # Upload to Google Sheets
        upload_to_google_sheets(df, client)
        
        print(f"\n{'='*60}")
        print("✅ PROCESS COMPLETED SUCCESSFULLY")
        print(f"{'='*60}")
        print(f"📁 Excel report saved as: {OUTPUT_FILE}")
        print(f"📊 Total rows processed: {len(df)}")
        print(f"🏢 Hubs processed: {len(HUBS)}")
        
        # Show hub column summary if available
        hub_column = find_hub_column(df)
        if hub_column and not df.empty:
            found_hubs = df[hub_column].unique()
            print(f"✅ Hubs found in report: {len(found_hubs)}")
        
        # Show extracted columns summary
        extracted_cols = ['Total Collection', 'Total Deposit', 'Overall Gap']
        available_cols = [col for col in extracted_cols if col in df.columns]
        print(f"📋 Columns extracted: {len(available_cols)}/{len(extracted_cols)}")
        for col in available_cols:
            print(f"   ✓ {col}")
        missing_cols = [col for col in extracted_cols if col not in df.columns]
        if missing_cols:
            print(f"⚠️ Missing columns:")
            for col in missing_cols:
                print(f"   ✗ {col}")
        
        print(f"\n📤 Data uploaded to Google Sheets:")
        print(f"   Spreadsheet: https://docs.google.com/spreadsheets/d/{OUTPUT_SPREADSHEET_ID}/edit")
        print(f"   Worksheet: {OUTPUT_WORKSHEET_NAME}")
        
        print(f"\n📅 Completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
    except Exception as e:
        print(f"\n{'='*60}")
        print("❌ ERROR OCCURRED")
        print(f"{'='*60}")
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()

